"""Stage 4 — write the finished leads out as CSV and XLSX."""

import csv
import logging
from pathlib import Path
from typing import Any, Dict, List

log = logging.getLogger(__name__)

COLUMNS: List[tuple[str, str]] = [
    ("title", "Objekt"),
    ("phone", "Telefon"),
    ("phone_source", "Telefon-Quelle"),
    ("email", "E-Mail"),
    ("website", "Website"),
    ("instagram", "Instagram"),
    ("host_name", "Gastgeber"),
    ("market", "Markt"),
    ("city", "Stadt"),
    ("country", "Land"),
    ("bedrooms", "Schlafzimmer"),
    ("person_capacity", "Gäste"),
    ("photo_count", "Fotos"),
    ("max_photo_width", "Max. Bildbreite"),
    ("rating", "Bewertung"),
    ("review_count", "Reviews"),
    ("is_superhost", "Superhost"),
    ("listings_by_host", "Objekte des Hosts"),
    ("url", "Airbnb-Link"),
]


def _rows(leads: List[Dict[str, Any]]) -> List[List[Any]]:
    return [[lead.get(key, "") for key, _ in COLUMNS] for lead in leads]


def to_csv(leads: List[Dict[str, Any]], path: Path) -> Path:
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh, delimiter=";")
        writer.writerow([label for _, label in COLUMNS])
        writer.writerows(_rows(leads))
    log.info("wrote %s (%s leads)", path, len(leads))
    return path


def to_xlsx(leads: List[Dict[str, Any]], path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_fill = PatternFill("solid", fgColor="0A0A0A")
    header_font = Font(bold=True, color="FFFFFF")
    for col, (_, label) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for row_idx, row in enumerate(_rows(leads), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col, (key, label) in enumerate(COLUMNS, start=1):
        widest = max([len(str(label))] + [len(str(lead.get(key, ""))) for lead in leads] or [10])
        ws.column_dimensions[get_column_letter(col)].width = min(max(widest + 2, 10), 55)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    log.info("wrote %s (%s leads)", path, len(leads))
    return path
